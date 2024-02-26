from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib
import os


root=Tk()
root.title("Veri Giriş Paneli")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")

file_path = os.path.join(os.getcwd(), 'Backened_data.xlsx')

if not os.path.exists(file_path):
    file = openpyxl.Workbook()
    sheet = file.active
    sheet['A1'] = "Tam İsim"
    sheet['B1'] = "Telefon Numarası"
    sheet['C1'] = "Yaş"
    sheet['D1'] = "Cinsiyet"
    sheet['E1'] = "Adres"

    file.save(file_path)

def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r'Backened_data.xlsx')

    messagebox.showinfo('Bilgilendirme','Kişi Eklendi!')


    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)



def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)

#İcon Bölümü
icon_image=PhotoImage(file="logo.png")
root.iconphoto(False, icon_image)


#Başlık Bölümü
Label(root,text="Lütfen bu giriş formunu doldurun:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#Label
Label(root,text="İsim",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="Telefon No",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="Yaş",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Cinsiyet",font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text="Adres",font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#girişler
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=25,bd=2,font=20)
contactEntry = Entry(root,textvariable=contactValue,width=25,bd=2,font=20)
ageEntry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

#gender
gender_combobox = Combobox(root,values=['Erkek','Kadın'],font='arial 14',state='r',width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('Erkek')

addressEntry = Text(root,width=50,height=4,bd=2)

nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

Button(root,text="Gönder",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Sil",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Çıkış",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)

root.mainloop()
